
from openpyxl import Workbook, load_workbook
import openpyxl
from selenium import webdriver
import time
import random
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import pyautogui as pag

def vao():
    b = entry1.get()
    path = Service(r"{}".format(b))
    cho = webdriver.ChromeOptions()
    cho.add_experimental_option("detach",True)
    dr = webdriver.Chrome(service=path,options=cho)
    dr.get('https://congdanhoctap.vn/dang-nhap')
    dr.maximize_window()
    return dr






def thd_ma():

    #vao chrome
    a = vao()
    wait = WebDriverWait(a,10)

    # lưu cccd bị lỗi
    data = []

    #vào đọc excel
    excel_path = entry2.get()
    wb = load_workbook(filename=excel_path)
    sheet = entry3.get()
    cot = entry4.get()
    ws = wb[sheet]
    co = ws[cot] 
    
    # tạo sheet excel mới để ghi các tk bị lỗi
    new_sheet =  wb.create_sheet(title='Sheet_dkbl')
    wb.save(filename=excel_path)
    # đăng nhập 
    def dang_nhap():
        for rep in range(0,len(co)):
            #Điền cccd
            val = co[rep].value
            dk = wait.until(EC.element_to_be_clickable((
                By.XPATH, '//*[@id="username"]'
                )))
            dk.send_keys('{}'.format(val))

            #Điền mk
            wait.until(EC.element_to_be_clickable((
            By.XPATH, '//*[@id="password"]'
            ))).send_keys('123456')    
            #bấm đăng nhập
            dang_nhap_button = a.find_elements(
                By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[3]/div/button'
                )
            dang_nhap_button1  = a.find_elements(
                By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[4]/div/button'
                )
            
            if dang_nhap_button :
                wait.until(EC.element_to_be_clickable((
                    By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[3]/div/button'
                    ))).click()
            elif dang_nhap_button1 :
                wait.until(EC.element_to_be_clickable((
                    By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[4]/div/button'
                    ))).click()        


        def kt_loi():
            er_ele = a.find_elements(By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[1]')
            achains = ActionChains(a)
            
            if er_ele:
                for i in er_ele:
                    er_ele_text = i.text

                if "Tài khoản hoặc mật khẩu không đúng" in er_ele_text:
                    
                    dk = wait.until(EC.element_to_be_clickable((
                            By.XPATH, '//*[@id="username"]'
                            )))
                    achains.move_to_element(dk).click().click().perform()
                    data.append(val)
                        
                elif "Khóa bảo vệ nội dung thông tin đã hết hạn" in er_ele_text:
                    a.refresh()
                    time.sleep(5)
                    loc = pag.locateOnScreen(image='tieptucproject.png')
                    time.sleep(5)
                    pag.click(loc)
                    time.sleep(2)
            else:
                #Thay Mã    
                ma = entry.get()
                ele = wait.until(EC.element_to_be_clickable((
                    By.XPATH, '//*[@id="unitcode"]'
                    ))) 
                achains.move_to_element(ele).click().click().click().send_keys('{}'.format(ma)).perform()
                #Lưu tt
                wait.until(EC.element_to_be_clickable((
                    By.XPATH, '/html/body/div/div[3]/div[2]/div[1]/form/div[5]/button'
                    ))).click()  
                #Đăng xuất
                wait.until(EC.element_to_be_clickable((
                    By.XPATH, '//*[@id="bs-example-navbar-collapse-1"]/ul[2]/li[2]/a'
                    ))).click()  
                
            kt_loi()
    #ghi tk lỗi vào sheet mới
    def ghi_moi():
        sheet_dkbl = wb['Sheet_dkbl']
        for re in data:
            sheet_dkbl.append([re])
        wb.save(filename=excel_path)

    ghi_moi()

