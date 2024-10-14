from openpyxl import Workbook, load_workbook
import openpyxl
from selenium import webdriver
import time
import random
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
def vao():
    path = Service(r"C:\Users\Administrator\Downloads\69696996\chromedriver.exe")
    cho = webdriver.ChromeOptions()
    cho.add_experimental_option("detach",True)
    dr = webdriver.Chrome(service=path,options=cho)
    dr.get('https://congdanhoctap.vn/dang-nhap')
    dr.maximize_window()
    return dr

def main():
    a = vao()
    wait = WebDriverWait(a,10)
    achains = ActionChains(a)

    wb = load_workbook('testexcel.xlsx')
    ws = wb['Sheet1']
    co = ws['A']
    


    wait.until(EC.element_to_be_clickable
                ((By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[3]/div/a[1]'))).click()
    wait.until(EC.element_to_be_clickable
                ((By.XPATH, '//*[@id="select2-addresscity-container"]'))).click()

    ele = a.find_elements(By.XPATH,
                            '//*[@id="addressdistrict"]')

    for v in range(len(co)):
            chon = co[v].value
            if ele:
                for i in ele:
                    txt = i.text.split('\n')
                    

    dic = {}
    for x in range(len(txt)):
            s = txt[x].strip()
            dic[s] = 2 + x
    print(dic)
"""                    
                if chon in txt:
                    wait.until(EC.element_to_be_clickable
                        ((By.XPATH, f'//*[@id="ethnic"]/option[{dic1[chon]}]'))).click()
                    print(dic1[chon])
                else:
                    print("Đéo thấy")"""


if __name__ == '__main__':
    main()