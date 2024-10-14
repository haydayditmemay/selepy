from selenium import webdriver
import time
import random
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pyautogui as pag
########################################################
#đọc excel
wb = load_workbook('Book2.xlsx')
ws = wb['Sheet1']
co = ws['B']
new_sheet =  wb.create_sheet(title='Sheet_dkbl')
wb.save('Book2.xlsx')

data = []

#vao chrome
def vao():
    path = Service(r"C:\Users\Administrator\Downloads\69696996\chromedriver.exe")
    cho = webdriver.ChromeOptions()
    cho.add_experimental_option("detach",True)
    dr = webdriver.Chrome(service=path,options=cho)
    dr.get('https://congdanhoctap.vn/dang-nhap')
    dr.maximize_window()
    return dr
a = vao()
wait = WebDriverWait(a,10)
achains = ActionChains(a)
for rep in range(1,len(co)):
    val = co[rep].value
    dk = wait.until(EC.element_to_be_clickable((
        By.XPATH, '//*[@id="username"]'
        )))
    dk.send_keys('{}'.format(val))
    

    wait.until(EC.element_to_be_clickable((
        By.XPATH, '//*[@id="password"]'
        ))).send_keys('123456') 
    dang_nhap_button = a.find_elements(By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[3]/div/button')
    dang_nhap_button1  = a.find_elements(By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[4]/div/button')
    if dang_nhap_button :
        wait.until(EC.element_to_be_clickable((
            By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[3]/div/button'
            ))).click()
    elif dang_nhap_button1 :
        wait.until(EC.element_to_be_clickable((
            By.XPATH, '/html/body/div/div[3]/div[2]/form/div/div/div/div[4]/div/button'
            ))).click()        


def ghi_moi():
    sheet_dkbl = wb['Sheet_dkbl']
    for re in data:
        sheet_dkbl.append([re])
    wb.save('Book2.xlsx')

ghi_moi()

